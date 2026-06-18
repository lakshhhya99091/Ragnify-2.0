"""
Excel exporters for bid extraction — produces two .xlsx files styled to match
the buyer's cost sheet:
  * Tender Details  — key/value table.
  * Asset Details   — Sr.No | Service Description | Qty | Sources + GRAND TOTAL.
"""
import os
import re
import sys
from io import BytesIO
from typing import Any, Dict

# Locally, openpyxl is installed under D:\Devtools\python-libs (the user keeps all
# downloaded packages off the C drive). On Render/Linux this path doesn't exist, so
# the guard is skipped and openpyxl is imported from the normal environment.
_DEVTOOLS_LIBS = r"D:\Devtools\python-libs"
if os.path.isdir(_DEVTOOLS_LIBS) and _DEVTOOLS_LIBS not in sys.path:
    sys.path.insert(0, _DEVTOOLS_LIBS)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from extractor import tender_rows

# ── Shared styles ───────────────────────────────────────────────────────────────
_BLUE = PatternFill("solid", fgColor="2D9CDB")        # title banner
_HEAD = PatternFill("solid", fgColor="DDEBF7")        # column headers
_YELLOW = PatternFill("solid", fgColor="FFF200")      # highlighted row
_ORANGE = PatternFill("solid", fgColor="F8CBAD")      # grand total
_WHITE_BOLD = Font(bold=True, color="FFFFFF", size=12)
_BOLD = Font(bold=True)
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(vertical="center", wrap_text=True)
_WRAP_TOP = Alignment(vertical="top", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="center")


def _safe_sheet_title(name: str) -> str:
    """Excel sheet titles: <=31 chars, none of : \\ / ? * [ ]."""
    return (re.sub(r'[:\\/?*\[\]]', "_", name) or "Sheet")[:31]


def _to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Tender Details ──────────────────────────────────────────────────────────────
def build_tender_xlsx(extraction: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tender Details"
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 55

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "Tender Details"
    c.fill = _BLUE
    c.font = _WHITE_BOLD
    c.alignment = _CENTER
    ws["B1"].border = _BORDER
    c.border = _BORDER

    r = 2
    for label, value in tender_rows(extraction):
        a, b = ws.cell(r, 1, label), ws.cell(r, 2, value)
        a.font = _BOLD
        a.alignment = _WRAP_TOP
        b.alignment = _WRAP_TOP
        a.border = b.border = _BORDER
        if label.startswith("No of technicians"):
            a.fill = b.fill = _YELLOW
        ws.row_dimensions[r].height = 28
        r += 1

    return _to_bytes(wb)


# ── Asset Details ───────────────────────────────────────────────────────────────
def build_asset_xlsx(extraction: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Details"
    widths = {"A": 8, "B": 46, "C": 10, "D": 50}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "List of computers and peripheral devices"
    t.fill = _BLUE
    t.font = _WHITE_BOLD
    t.alignment = _CENTER
    for col in range(1, 5):
        ws.cell(1, col).border = _BORDER

    headers = ["Sr.No.", "Service Description", "Qty", "Sources"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(2, col, h)
        cell.font = _BOLD
        cell.fill = _HEAD
        cell.alignment = _CENTER
        cell.border = _BORDER

    assets = (extraction or {}).get("asset_details", []) or []
    r = 3
    total = 0
    for i, item in enumerate(assets, start=1):
        qty = int(item.get("quantity", 0) or 0)
        total += qty
        srcs = item.get("sources", []) or []
        src_text = " | ".join(f"{s.get('source','?')}: {s.get('quantity',0)}" for s in srcs)
        row = [i, item.get("description", ""), qty, src_text]
        for col, val in enumerate(row, start=1):
            cell = ws.cell(r, col, val)
            cell.border = _BORDER
            cell.alignment = _CENTER if col in (1, 3) else _WRAP
        ws.row_dimensions[r].height = 30
        r += 1

    if not assets:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
        cell = ws.cell(3, 1, "No asset details were found in the document or its accessible linked sources.")
        cell.alignment = _WRAP
        cell.border = _BORDER
        r = 4

    # GRAND TOTAL row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    gt = ws.cell(r, 1, "GRAND TOTAL")
    gt.font = _BOLD
    gt.alignment = _CENTER
    tot = ws.cell(r, 3, total)
    tot.font = _BOLD
    tot.alignment = _CENTER
    for col in range(1, 5):
        cell = ws.cell(r, col)
        cell.fill = _ORANGE
        cell.border = _BORDER

    return _to_bytes(wb)


def safe_filename(base: str, suffix: str) -> str:
    """Build a download filename from the document name."""
    stem = os.path.splitext(base or "document")[0]
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("_") or "document"
    return f"{stem}_{suffix}.xlsx"
